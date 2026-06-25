"""
build.py — Rebuild the dashboard with the latest data from Excel files.

Run this every time you update data.xlsx or Used_Cars_Spends.xlsx:
    python build.py

It reads both Excel files, processes them, and embeds the result
directly into index.html. No internet, no upload, no GitHub needed.
"""

import openpyxl, json, re, sys
from datetime import datetime
from pathlib import Path

HERE = Path(__file__).parent

CAMP_MAP = {
    'AHM-Search-Buy-Used-Car-8april':   {'city': 'Ahmedabad',  'platform': 'Google'},
    'CHD-Buyer-Acquisition-6April':     {'city': 'Chandigarh', 'platform': 'Google'},
    'Nashik-Search-Buy-UsedCar-8april': {'city': 'Nashik',     'platform': 'Google'},
    'Nasik-Search-Buy-UsedCar-8april':  {'city': 'Nashik',     'platform': 'Google'},
    'AHM-UCR-Catalogue-Ads':            {'city': 'Ahmedabad',  'platform': 'Meta'},
    'Chandigarh-UCR-Catalogue-Ads':     {'city': 'Chandigarh', 'platform': 'Meta'},
    'Nashik-UCR-Catalogue-Ads':         {'city': 'Nashik',     'platform': 'Meta'},
}
MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

def to_date_str(d):
    return d.strftime('%Y-%m-%d')

def get_month(d):
    return f"{MONTHS[d.month-1]}'{str(d.year)[2:]}"

def wom(ds):
    return (int(ds[8:10]) - 1) // 7 + 1

def detect_city(camp):
    if 'AHM' in camp: return 'Ahmedabad'
    if 'CHD' in camp or 'Chandigarh' in camp: return 'Chandigarh'
    if 'Nashik' in camp or 'Nasik' in camp: return 'Nashik'
    return None

def build():
    leads_path  = HERE / 'data.xlsx'
    spends_path = HERE / 'Used_Cars_Spends.xlsx'
    html_path   = HERE / 'index.html'

    for p in [leads_path, spends_path, html_path]:
        if not p.exists():
            print(f"ERROR: {p.name} not found in {HERE}")
            sys.exit(1)

    merged = {}

    # ── Leads (data.xlsx) ────────────────────────────────────────────────────
    print(f"Reading {leads_path.name}...")
    wb = openpyxl.load_workbook(leads_path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(header, row))
        dt = d.get('Date')
        if not isinstance(dt, datetime): continue
        rc = str(d.get('utm_campaign') or '').strip()
        camp = 'Nashik-Search-Buy-UsedCar-8april' if rc == 'Nasik-Search-Buy-UsedCar-8april' else rc
        if not camp: continue
        city = str(d.get('CITY') or '').strip()
        if city not in ('Ahmedabad', 'Chandigarh', 'Nashik'): continue
        cm = CAMP_MAP.get(camp)
        plat = cm['platform'] if cm else ('Meta' if 'UCR' in camp or 'Catalogue' in camp else 'Google')
        ds, mo, w = to_date_str(dt), get_month(dt), wom(to_date_str(dt))
        key = f"{ds}|{camp}|{city}|{mo}|{w}|{plat}"
        if key not in merged:
            merged[key] = {'dateStr': ds, 'campaign': camp, 'city': city, 'month': mo,
                           'wom': w, 'Platform': plat, 'spends': 0,
                           'gen_leads': 0, 'triggered': 0, 'dealer_triggers': 0}
        merged[key]['gen_leads']  += int(d.get('Total_lead') or 0)
        merged[key]['triggered']  += int(d.get('Unique') or 0)
        if d.get('Listing_Group') == 'Dealer':
            merged[key]['dealer_triggers'] += int(d.get('Unique') or 0)

    # ── Spends (Used_Cars_Spends.xlsx) ───────────────────────────────────────
    print(f"Reading {spends_path.name}...")
    wb2 = openpyxl.load_workbook(spends_path, data_only=True)

    for sheet_name, plat_default in [('Ga', 'Google'), ('FB', 'Meta')]:
        if sheet_name not in wb2.sheetnames:
            continue
        ws2 = wb2[sheet_name]
        h = [c.value for c in ws2[1]]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            r = dict(zip(h, row))
            if sheet_name == 'FB':
                raw_date = r.get('Reporting starts') if r.get('Reporting starts') is not None else r.get('Date')
                raw_camp = r.get('Campaign name') if r.get('Campaign name') is not None else r.get('Campaign Name')
                raw_amt  = r.get('Amount spent (INR)') if r.get('Amount spent (INR)') is not None else r.get('Amount Spent')
            else:
                raw_date, raw_camp, raw_amt = r.get('Day'), r.get('Campaign'), r.get('Cost')
            amt = float(raw_amt or 0)
            if not amt > 0: continue
            if not isinstance(raw_date, datetime): continue
            camp = str(raw_camp or '').strip()
            if not camp: continue
            cm = CAMP_MAP.get(camp)
            city = cm['city'] if cm else detect_city(camp)
            if not city: continue
            plat = cm['platform'] if cm else plat_default
            ds, mo, w = to_date_str(raw_date), get_month(raw_date), wom(to_date_str(raw_date))
            key = f"{ds}|{camp}|{city}|{mo}|{w}|{plat}"
            if key in merged:
                merged[key]['spends'] = round((merged[key]['spends'] + amt) * 100) / 100
            else:
                merged[key] = {'dateStr': ds, 'campaign': camp, 'city': city, 'month': mo,
                               'wom': w, 'Platform': plat,
                               'spends': round(amt * 100) / 100,
                               'gen_leads': 0, 'triggered': 0, 'dealer_triggers': 0}

    data = sorted(merged.values(), key=lambda r: (r['dateStr'], r['city']))
    all_dates = sorted(set(r['dateStr'] for r in data))
    latest = all_dates[-1] if all_dates else '—'
    print(f"Processed {len(data)} rows | {all_dates[0]} → {latest}")

    # ── Embed into HTML ──────────────────────────────────────────────────────
    print(f"Updating {html_path.name}...")
    html = html_path.read_text(encoding='utf-8')

    new_data_js = 'var DATA = ' + json.dumps(data, separators=(',', ':')) + ';'

    old_start = html.find('var DATA = [')
    old_end   = html.find('];', old_start) + 2
    if old_start == -1:
        print("ERROR: Could not find 'var DATA = [' in index.html")
        sys.exit(1)

    html_new = html[:old_start] + new_data_js + html[old_end:]
    html_path.write_text(html_new, encoding='utf-8')

    # Format latest date nicely
    parts = latest.split('-')
    mn = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    fmt = f"{int(parts[2])} {mn[int(parts[1])-1]} {parts[0]}" if len(parts)==3 else latest

    print(f"\n✓ Done! Dashboard updated — data up to {fmt}")
    print(f"  Open index.html in your browser or push to GitHub to publish.")

if __name__ == '__main__':
    build()
